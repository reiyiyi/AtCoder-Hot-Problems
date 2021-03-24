from flask import Flask, render_template
import boto3
import openpyxl
import os
import datetime
import aws
import filename

app = Flask(__name__)

@app.route('/')
def index():
    file_name = filename.get()

    s3 = aws.get()

    bucket = s3.Bucket('hotproblems')
    bucket.download_file('hot_problems_data/' + file_name, file_name)

    wb = openpyxl.load_workbook(file_name)
    sheet = wb.worksheets[0]

    problems = [[sheet.cell(row = i+1, column = j+1).value for j in range(5)] for i in range(100)] #jsonにする

    date = datetime.datetime.now() + datetime.timedelta(hours=9) - datetime.timedelta(days=1)
    date = str(date.year) + "年" + str(date.month) + "月" + str(date.day) + "日"

    start_time = datetime.datetime.fromtimestamp(int(sheet.cell(row = 1, column = 6).value))
    end_time = datetime.datetime.fromtimestamp(int(sheet.cell(row = 2, column = 6).value))
    code_num = sheet.cell(row = 3, column = 6).value

    return render_template('index.html',
                        problems=problems,
                        date=date,
                        start_time=start_time,
                        end_time=end_time,
                        code_num=code_num) #problemsを配列からjson(dict)にしたい

if __name__ == "__main__":
   app.run(debug=True,host='0.0.0.0')

#http://localhost:5000