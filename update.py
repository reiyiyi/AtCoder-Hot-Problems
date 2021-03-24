import requests
import json
import time
import datetime
import openpyxl
import boto3
import os
import aws
import filename

file_name = filename.get()
wb = openpyxl.Workbook()
sheet = wb.worksheets[0]

s3 = aws.get()

bucket = s3.Bucket('hotproblems')

current_time = int(time.time()) + 9 * 60 * 60
research_time = current_time - 24 * 60 * 60

sheet.cell(row = 1, column = 6).value = research_time
sheet.cell(row = 2, column = 6).value = current_time

submissions_url = "https://kenkoooo.com/atcoder/atcoder-api/v3/from/"
contests_url = "https://kenkoooo.com/atcoder/resources/contests.json"
problems_url = "https://kenkoooo.com/atcoder/resources/problems.json"

contests_response = requests.get(contests_url)
print("hit API...")
time.sleep(1)
problems_response = requests.get(problems_url)
print("hit API...")
time.sleep(1)

contests_json_data = contests_response.json()
problems_json_data = problems_response.json()

problems_count = dict()
contests_name = dict()

for data in contests_json_data:
    contests_name[data["id"]] = data["title"]

for data in problems_json_data:
    problems_count[data["id"]] = [0, contests_name[data["contest_id"]], data["title"], data["contest_id"], data["id"]]

num = 0

for _ in range(500):
    submissions_response = requests.get(submissions_url + str(research_time))
    print("hit API...")
    time.sleep(1)
    submissions_json_data = submissions_response.json()

    for data in submissions_json_data:
        if data["epoch_second"] >= current_time:
            research_time = current_time
            break

        if num == 0:
            print(datetime.datetime.fromtimestamp(int(data["epoch_second"])))

        problems_count[data["problem_id"]][0] += 1
        num += 1
    
    if len(submissions_json_data) == 0:
        break

    research_time = submissions_json_data[-1]["epoch_second"] + 1

    if data["epoch_second"] >= current_time:
        break

sheet.cell(row = 3, column = 6).value = num

problems_count = sorted(problems_count.items(), key=lambda x: x[1][0], reverse=True)
problems_count = problems_count[0:100]

for i in range(100):
    sheet.cell(row = i+1, column = 1).value = problems_count[i][1][1]
    sheet.cell(row = i+1, column = 2).value = problems_count[i][1][2]
    sheet.cell(row = i+1, column = 3).value = problems_count[i][1][0]
    sheet.cell(row = i+1, column = 4).value = problems_count[i][1][3]
    sheet.cell(row = i+1, column = 5).value = problems_count[i][1][4] #jsonにする

wb.save(file_name)

bucket.upload_file(file_name, 'hot_problems_data/' + file_name)